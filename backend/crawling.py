#-*- coding:utf-8 -*-
from selenium import webdriver
from bs4 import BeautifulSoup
from time import sleep
from openpyxl import Workbook,load_workbook

from selenium.webdriver.common.by import By

# Chrome의 경우 | 아까 받은 chromedriver의 위치를 지정해준다.
from selenium.webdriver.chrome.options import Options
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)
# driver = webdriver.Chrome('./chromedriver.exe')

# url에 접근한다.
driver.get('https://everytime.kr/login')
# 암묵적으로 웹 자원 로드를 위해 5초까지 기다려 준다.
driver.implicitly_wait(5)

# 아이디/비밀번호를 입력해준다.
driver.find_element("name", 'id').send_keys('skybro2004')
driver.find_element("name", 'password').send_keys('jun040309')

# 로그인 버튼을 눌러주자.
driver.find_element("xpath", '/html/body/div[1]/div/form/input').click()

input("Press ENTER to proceed")

driver.get('https://everytime.kr/timetable')

#수업 목록에서 검색 클릭
driver.find_element("xpath", '//*[@id="container"]/ul/li[1]').click()

input("Press ENTER to proceed")

#팝업창 닫기
sleep(2)
# driver.find_element("xpath", '//*[@id="sheet"]/ul/li[3]/a').click()

input("Press ENTER to proceed")

pre_count = 0
#스크롤 맨아래로 내리기
while True:
    #tr요소 접근
    element = driver.find_elements(By.CSS_SELECTOR, "#subjects > div.list > table > tbody > tr")

    # tr 마지막 요소 접근
    result = element[-1]
    #마지막요소에 focus주기
    driver.execute_script('arguments[0].scrollIntoView(true);',result)
    sleep(5)

    #현재 접근한 요소의 갯수
    current_count = len(element)
    if pre_count == current_count:
        break
    #같지않다면
    pre_count = current_count


html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')

trs = soup.select('#subjects > div.list > table > tbody > tr')

results = []

for tr in trs:
    result=[]
    tds = tr.select('#subjects > div.list > table > tbody > tr > td')
    # result.append(tds[0].text) #강의계획서 링크
    result.append(tds[1].text) #개설학년
    result.append(tds[2].text) #이수구분(전공/교양)
    result.append(tds[3].text) #학수번호
    result.append(tds[4].text) #과목명
    result.append(tds[5].text) #교수명
    result.append(tds[6].text) #학점
    result.append(tds[7].text) #시간
    result.append(tds[8].text) #강의실
    temp = tds[9].select('a')[0]['title']
    result.append(temp) #강의평
    result.append(tds[10].text) #담은인원
    result.append(tds[11].text) #정원
    result.append(tds[12].text) #개설학과
    result.append(tds[13].text) #비고
    results.append(result)
    # print(result)
#값이 들어있다면!
if results:
    print("성공!!")
excel_column = 9
write_wb = Workbook()
write_ws = write_wb.create_sheet('result.xls')
for data in results:
    write_ws = write_wb.active
    write_ws.append(data)
write_wb.save('C:/Users/skybr/Downloads/everytime-timetable-crawling-master/everytime-timetable-crawling-master/asdf.csv')



# Naver 페이 들어가기
# driver.get('https://order.pay.naver.com/home')
# html = driver.page_source
# soup = BeautifulSoup(html, 'html.parser')
# notices = soup.select('div.goods_item > div > a > p')
#
# for n in notices:
#     print(n.text.strip())
